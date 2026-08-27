# pip install openpyxl
# python check_batpac.py

import openpyxl

wb_f = openpyxl.load_workbook(
    r'C:\Users\ishaa\OneDrive\College Syllabus & Assignments\Term 3\Work\BatPaC 5.2 2024-11-1 - Copy.xlsm',
    data_only=False
)
wb_v = openpyxl.load_workbook(
    r'C:\Users\ishaa\OneDrive\College Syllabus & Assignments\Term 3\Work\BatPaC 5.2 2024-11-1 - Copy.xlsm',
    data_only=True
)

sheet_f = wb_f['Battery Design']
sheet_v = wb_v['Battery Design']

# These are the columns that contain real calculated values for Battery 1
# G = Battery 1, B = row description
CALC_COLS = ['G']
LABEL_COL = 'B'

# Formulas that are just lookups/boilerplate - skip these
SKIP_PATTERNS = ['INDEX(', 'IFERROR(IF(', 'IF(AND(D', 'IF(INDEX(']

def is_boilerplate(formula):
    if formula is None:
        return True
    for pattern in SKIP_PATTERNS:
        if pattern in str(formula):
            return True
    return False

with open('battery_design_sheet.txt', 'w', encoding='utf-8') as out:
    out.write("=" * 70 + "\n")
    out.write("BatPaC v5.2 — Battery Design Sheet (Battery 1, Column G)\n")
    out.write("Only showing real calculated formulas, not lookup boilerplate\n")
    out.write("=" * 70 + "\n\n")

    current_section = ""

    for row in range(1, sheet_f.max_row + 1):
        label = sheet_f[f'{LABEL_COL}{row}'].value

        # Detect section headers (cells in col B with no formula in col G)
        g_formula = sheet_f[f'G{row}'].value
        if label and g_formula is None and isinstance(label, str) and len(label) > 3:
            current_section = label.strip()
            out.write(f"\n{'─' * 70}\n")
            out.write(f"SECTION: {current_section}\n")
            out.write(f"{'─' * 70}\n")
            continue

        for col in CALC_COLS:
            cell_ref = f'{col}{row}'
            formula = sheet_f[cell_ref].value
            value   = sheet_v[cell_ref].value

            if formula is None and value is None:
                continue
            if is_boilerplate(formula):
                continue
            if formula == value:  # plain hardcoded number, not a formula
                continue

            label_clean = str(label).strip() if label else "—"

            out.write(f"\nRow {row:>4} | {cell_ref}\n")
            out.write(f"  Description : {label_clean}\n")
            out.write(f"  Formula     : {formula}\n")
            out.write(f"  Value       : {value}\n")

print("Done - open battery_design_sheet.txt")